import time
import os
from datetime import datetime
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# === Configuration ===
DOWNLOAD_PATH = r"C:\Users\ksaur\Downloads"
XLSX_FILE = os.path.join(DOWNLOAD_PATH, "serp_rank_history.xlsx")
TARGET_SITE = "vypzee.com"
KEYWORDS = [
    "chawri Bazar",
    "karol bagh shops",
    "chandni chowk shops",
    "M3M Urbana shops",
    "M3M 65th Avenue shops",
    "M3M IFC shops"
]
MAX_PAGES = 5

# === Google SERP Rank Fetcher ===
def get_google_rank(driver, keyword, target_site, max_pages=5):
    keyword_encoded = keyword.replace(" ", "+")
    base_url = f"https://www.google.com/search?q={keyword_encoded}"

    for page in range(max_pages):
        url = base_url + f"&start={page * 10}"
        driver.get(url)
        time.sleep(2)
        results = driver.find_elements(By.CSS_SELECTOR, 'div.MjjYud')
        for idx, result in enumerate(results):
            links = result.find_elements(By.TAG_NAME, 'a')
            for link in links:
                href = link.get_attribute('href')
                if href and target_site in href:
                    return page * 10 + idx + 1
    return None

# === Excel Updater with Color ===
def update_workbook_with_data(today_data, date_str):
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        sheet_history = wb["History"]
        sheet_insights = wb["Insights"]
    else:
        wb = Workbook()
        sheet_history = wb.active
        sheet_history.title = "History"
        sheet_history.append(["Keyword", "Rank", "Date"])
        sheet_insights = wb.create_sheet("Insights")
        sheet_insights.append(["Keyword"])  # First column is Keyword

    # === Update History Sheet ===
    for keyword, rank in today_data.items():
        sheet_history.append([
            keyword,
            rank if rank else f"Not in Top {MAX_PAGES * 10}",
            date_str
        ])

    # === Update Insights Sheet ===
    # Add date column if not already present
    if date_str not in [cell.value for cell in sheet_insights[1]]:
        sheet_insights.cell(row=1, column=sheet_insights.max_column + 1).value = date_str

    # Get the column index for today's date
    date_col = None
    for idx, cell in enumerate(sheet_insights[1], start=1):
        if cell.value == date_str:
            date_col = idx
            break

    # Map keywords to their row numbers
    keyword_rows = {sheet_insights.cell(row=r, column=1).value: r for r in range(2, sheet_insights.max_row + 1)}

    for keyword, rank in today_data.items():
        rank_value = rank if rank else f"Not in Top {MAX_PAGES * 10}"

        if keyword in keyword_rows:
            row = keyword_rows[keyword]
        else:
            row = sheet_insights.max_row + 1
            sheet_insights.cell(row=row, column=1).value = keyword

        cell = sheet_insights.cell(row=row, column=date_col)
        cell.value = rank_value

        # Apply color formatting
        if isinstance(rank, int):
            if rank <= 5:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            elif rank <= 10:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow
            else:
                cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Red
        else:
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Red

    # Make headers bold
    for cell in sheet_insights[1]:
        cell.font = Font(bold=True)
    for cell in sheet_history[1]:
        cell.font = Font(bold=True)

    wb.save(XLSX_FILE)

# === Main Execution ===
if __name__ == "__main__":
    today_str = datetime.now().strftime("%Y-%m-%d")
    options = uc.ChromeOptions()
    # options.add_argument("--headless")  # Uncomment to run in background
    driver = uc.Chrome(options=options)

    try:
        today_ranks = {}
        print("\n🔍 Checking SERP Ranks...\n")

        for keyword in KEYWORDS:
            rank = get_google_rank(driver, keyword, TARGET_SITE, MAX_PAGES)
            rank_display = rank if rank else f"Not in Top {MAX_PAGES * 10}"
            print(f"✅ '{keyword}' → Rank: {rank_display}")
            today_ranks[keyword] = rank

        update_workbook_with_data(today_ranks, today_str)
        print(f"\n📊 Excel updated with today's ranks: {XLSX_FILE}")

    finally:
        driver.quit()
